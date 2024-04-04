import pandas as pd
import datetime as dt
import pytz
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side


noww = dt.datetime.now(pytz.timezone("Asia/Bangkok")).date().strftime(format="%Y%m%d")
toDate = dt.datetime.now(pytz.timezone("Asia/Bangkok")).date().strftime(format="%d/%m/%Y")
fromDate = (dt.datetime.now(pytz.timezone("Asia/Bangkok")).date() - dt.timedelta(days=7)).strftime(format="%d/%m/%Y")
dailyFolder = r"/home/nhdminh/airflow/DMS_daily"
reportFolder = os.path.join(dailyFolder, noww)
row_start = 7
col_start = 2


# df = []
# for f in os.listdir(reportFolder):
#     df.append(pd.read_excel(os.path.join(reportFolder, f)))

# with pd.ExcelWriter(os.path.join(dailyFolder, "TestMultiSheet.xlsx")) as writer:
#     # use to_excel function and specify the sheet_name and index
#     # to store the dataframe in specified sheet
#     for row in range(0, len(df)):
#         print(row)
#         df[row].to_excel(writer, sheet_name=str(row), index=False)


# Bao_cao_chi_tiet_VTKHC2_20231209083001.xlsx
fileName = os.path.join(reportFolder, "Bao_cao_chi_tiet_VTKHC2_20231209083001.xlsx")
df = pd.read_excel(fileName)
df = df[3:].reset_index().drop(columns={"index"})
df.columns = df.loc[0]
df = df.loc[1:]
df = df.drop(columns="STT")


wbName = os.path.join(dailyFolder, "TestMultiSheet.xlsx")
wb = openpyxl.load_workbook(wbName)
print(wb.sheetnames)
sheet = wb["0"]

if "99" in wb.sheetnames:
    print(wb.sheetnames)
    wb.remove(wb["99"])

wb.create_sheet(
    index=9,
    title="99",
)


ws = wb["99"]


def dataframe_to_sheet(df, ws, row_i, col_i):
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + row_i, column=c_idx + col_i, value=value)
    ws.delete_rows(row_i + 2, 1)


dataframe_to_sheet(df, ws, row_start, col_start)

for column in range(1, ws.max_column + 1):
    col_letter = get_column_letter(column)
    ws.column_dimensions[col_letter].bestFit = True
print()
print(get_column_letter(ws.max_column))
print(str(len(df)))


def set_border(ws, cell_range, border):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)

sheet_start = get_column_letter(col_start + 1) + str(row_start + 1)
sheet_end = get_column_letter(ws.max_column) + str(row_start + 1 + len(df))
cell_range = ws[sheet_start:sheet_end]


print(f"{sheet_start}:{sheet_end}")
set_border(ws, f"{sheet_start}:{sheet_end}", thin_border)

ws["A1"] = "Báo cáo DMS"
ws["A2"] = "Công ty vùng miền"
ws["A3"] = "Từ ngày đến ngày"
ws["A4"] = "Nội dung báo cáo"
# >>> cell_range = ws['A1':'C2']


if "22" in wb.sheetnames:
    print(wb.sheetnames)
    wb.remove(wb["22"])

wb.create_sheet(
    index=9,
    title="22",
)
ws22 = wb["22"]
ws22["A1"] = "Báo cáo DMS"
ws22["A2"] = "Công ty vùng miền"
ws22["A3"] = "Từ ngày đến ngày"
ws22["A4"] = "Nội dung báo cáo"
# >>> cell_range = ws['A1':'C2']
wb.save(wbName)
