# %%
import datetime as dt
import os
import shutil
import pytz

import numpy as np
import openpyxl
import pandas as pd
import xlsxwriter
import xlwings as xw
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

print(sys.getdefaultencoding())
print(os.popen("taskkill /f /t /im EXCEL.exe").read())
print(os.popen("Taskkill /f /im onedrive.exe").read())

# %%
# Get reported date

noww = dt.datetime.now(pytz.timezone("Asia/Bangkok")).date().strftime(format="%Y%m%d")
toDate = dt.datetime.now(pytz.timezone("Asia/Bangkok")).date().strftime(format="%d/%m/%Y")
fromDate = (dt.datetime.now(pytz.timezone("Asia/Bangkok")).date() - dt.timedelta(days=7)).strftime(format="%d/%m/%Y")
parent_path = os.path.abspath(os.path.join(os.path.abspath(""), os.pardir))
dailyFolder = f"/home/nhdminh/DMS/DMS_daily"
print(dailyFolder)
reportFolder = os.path.abspath(os.path.join(dailyFolder, noww))
print(reportFolder)
row_start = 5
col_start = 0


# %%
def DMStoExcel(dfExport, file_name, sheet_name, index):
    # exporting pandas DataFrames to Excel files.
    # Check if the Excel file already exists
    if os.path.exists(file_name):
        try:
            # If the file exists, open it in 'append' mode and replace the sheet if it already exists
            with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                # Export the DataFrame to the specified sheet in the Excel file
                dfExport.to_excel(writer, sheet_name=sheet_name, index=index)

        except Exception as error:
            # If there's an error during the export, print the error message and continue
            print(error)
            pass
    else:
        # If the Excel file doesn't exist, create it
        with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
            # Export the DataFrame to the specified sheet in the Excel file
            dfExport.to_excel(writer, sheet_name=sheet_name, index=index)


# %%
def DMStoExcelRange(Worksheet, dfExport, range):
    Worksheet.range(range).options(index=False).value = dfExport
    range = Worksheet.range(range).expand()
    range.api.Borders.Weight = 2
    range.autofit()


# %%


def dataframe_to_sheet(df, ws, row_i, col_i):
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, 0):
        for c_idx, value in enumerate(row, 1):
            # print(f"{c_idx}, {value}")
            if c_idx == 1:
                continue
            ws.cell(row=r_idx + row_i, column=c_idx + col_i - 1, value=value)
    ws.delete_rows(row_i + 1, 1)


def set_border(ws, cell_range, border):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


# %% BCGheThamC2
def BCGheThamC2(dfExport, file_name, sheet_name, reportFolder):
    #######################
    dfExport.loc[:, ["Tỉnh/thành phố"]].dropna()
    dfExport.loc[:, ["MÃ NVTT"]].dropna()
    dfExport.loc[:, ["Ngày"]].dropna()
    dfExport = dfExport.reset_index(drop=True)
    dfExport.index = dfExport.index + 1
    # DMStoExcel(dfExport, file_name, sheet_name, False)
    # %% Save báo cáo tổng hợp
    noLoaiKH = dfExport["Loại KH"].drop_duplicates().count()
    LoaiKH = dfExport["Loại KH"].drop_duplicates()
    dfExport = dfExport.loc[:, ["Miền", "MÃ NVTT", "TÊN NVTT", "Mã KH", "Tỉnh/thành phố", "Ngày", "Loại KH"]]

    dfExport_pivot = pd.pivot_table(
        dfExport,
        values="Mã KH",
        index=["MÃ NVTT", "TÊN NVTT", "Tỉnh/thành phố", "Ngày"],
        columns=["Loại KH"],
        aggfunc="count",
        fill_value=0,
    )
    dfExport_pivot = dfExport_pivot.reset_index()
    dfExport_pivot["Tổng cộng"] = dfExport_pivot.apply(lambda x: np.sum([x[lkh] for lkh in LoaiKH]), axis=1)

    try:
        dfExport_group = dfExport_pivot.groupby("TÊN NVTT").sum(numeric_only=True).reset_index()
        dfExport_group.loc[:, "MÃ NVTT"] = "zTổng cộng"
    except Exception as error:
        # If there's an error during the export, print the error message and continue
        print(error)
        pass
    Total = []
    for i in range(1, noLoaiKH + 2):
        Total.append(int(dfExport_pivot.iloc[0:, [-i]].sum().iloc[0]))
        # Total.append(int(dfExport_pivot.iloc[0:, [-i]].sum()))

    dfExport_pivot = (
        pd.concat([dfExport_pivot, dfExport_group])
        .sort_values(by=["TÊN NVTT", "MÃ NVTT"], ascending=True)
        .reset_index(drop=True)
    )

    filter = dfExport_pivot["MÃ NVTT"] == "zTổng cộng"
    dfExport_pivot.loc[filter, "MÃ NVTT"] = "Tổng cộng     "
    dfExport_pivot.index = dfExport_pivot.index + 1
    dfExport_pivot = dfExport_pivot.reset_index()
    dfExport_pivot = dfExport_pivot.rename(columns={"index": "STT"})
    # %% Save to workbook
    #
    try:
        Workbook = openpyxl.load_workbook(file_name)
    except Exception as e:
        print(e)
        Workbook = openpyxl.Workbook()
        Workbook.save(file_name)
        print(f"Create file {file_name}")

    if sheet_name in Workbook.sheetnames:
        Workbook.remove(Workbook[sheet_name])

    Workbook.create_sheet(
        index=0,
        title=sheet_name,
    )
    Worksheet = Workbook[sheet_name]
    # %% Insert các thông tin chung
    #

    Worksheet["A1"].value = "Báo cáo ghé thăm C2 trong tuần"
    Worksheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
    Worksheet["A1"].font = Font(bold=True)
    # ws.merge_cells(start_row=startcell, start_column=1, end_row=startcell+mergecount, end_column=1)
    Worksheet["A2"] = dfExport["Miền"].loc[dfExport.index[0]]
    Worksheet.merge_cells(start_row=2, end_row=2, start_column=1, end_column=6)

    Worksheet["A3"] = "Từ ngày " + fromDate + " đến hết ngày " + toDate
    Worksheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=6)

    dataframe_to_sheet(dfExport_pivot, Worksheet, row_start, col_start)

    print(f"Save BCGheThamC2 {file_name} => {sheet_name}")
    # %% Edit table pivot
    #
    sheet_start = get_column_letter(col_start + 1) + str(row_start)
    sheet_end = get_column_letter(Worksheet.max_column) + str(row_start + 1 + len(dfExport_pivot))

    # print(f"{sheet_start}:{sheet_end}")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )
    set_border(Worksheet, f"{sheet_start}:{sheet_end}", thin_border)

    for i in range(1, Worksheet.max_column + 1):
        header = get_column_letter(i) + str(row_start)
        Worksheet[header].font = Font(bold=True)

    # bold dòng tổng cộng
    for row in range(row_start, len(dfExport_pivot) + row_start + 1):
        if str(Worksheet["D" + str(row)].value) == "nan":
            for i in range(1, Worksheet.max_column + 1):
                header = get_column_letter(i) + str(row)
                Worksheet[header].font = Font(bold=True)
    Workbook.save(file_name)


# %% BCDonHangC2
def BCDonHangC2(dfExport, file_name, sheet_name, reportFolder):
    # %% Save báo cáo tổng hợp
    dfExport = dfExport.loc[
        :,
        [
            "Miền",
            "Mã NVTT",
            "Tên NVTT",
            "Mã KH",
            "Tên KH",
            "Số đơn hàng",
            "Mã SP",
            "Tên Sản phẩm",
            "Số lượng (tấn)",
            "Thành tiền (VND)",
        ],
    ]
    dfExport_pivot = pd.pivot_table(
        dfExport,
        values=["Số lượng (tấn)", "Thành tiền (VND)"],
        index=[
            "Miền",
            "Mã NVTT",
            "Tên NVTT",
            "Mã KH",
            "Tên KH",
            "Số đơn hàng",
            "Mã SP",
            "Tên Sản phẩm",
        ],
        aggfunc="sum",
        fill_value=0,
    )
    dfExport_pivot = dfExport_pivot.reset_index()
    dfExport_group = dfExport_pivot.groupby(["Tên NVTT"]).sum(numeric_only=True).reset_index()

    try:
        dfExport_group.loc[:, "Miền"] = "zTổng cộng"
    except Exception as error:
        # If there's an error during the export, print the error message and continue
        print(error)
        pass
    dfExport_pivot = (
        pd.concat([dfExport_pivot, dfExport_group])
        .sort_values(by=["Tên NVTT", "Miền"], ascending=True)
        .reset_index(drop=True)
    )

    dfExport_pivot.index = dfExport_pivot.index + 1
    dfExport_pivot = dfExport_pivot.reset_index()
    dfExport_pivot = dfExport_pivot.rename(columns={"index": "STT"})
    dfExport_pivot = dfExport_pivot.reset_index(drop=True)

    filter = dfExport_pivot["Miền"] == "zTổng cộng"
    dfExport_pivot.loc[filter, "Miền"] = "Tổng cộng     "

    # %% Save to excel
    try:
        Workbook = openpyxl.load_workbook(file_name)
    except:
        Workbook = openpyxl.Workbook()
        Workbook.save(file_name)
        print(f"Create file {file_name}")

    if sheet_name in Workbook.sheetnames:
        Workbook.remove(Workbook[sheet_name])

    Workbook.create_sheet(
        index=0,
        title=sheet_name,
    )
    Worksheet = Workbook[sheet_name]

    # %%

    Worksheet["A1"].value = "Báo cáo đơn hàng C2 trong tuần"
    Worksheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
    Worksheet["A1"].font = Font(bold=True)
    Worksheet["A2"] = dfExport["Miền"].loc[dfExport.index[0]]
    Worksheet.merge_cells(start_row=2, end_row=2, start_column=1, end_column=6)

    Worksheet["A3"] = "Từ ngày " + fromDate + " đến hết ngày " + toDate
    Worksheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=6)

    dataframe_to_sheet(dfExport_pivot, Worksheet, row_start, col_start)
    TotalSL = dfExport_pivot["Số lượng (tấn)"].sum() / 2
    TotalTT = dfExport_pivot["Thành tiền (VND)"].sum() / 2

    # dfExport_pivot["Số lượng (tấn)"] = dfExport_pivot["Số lượng (tấn)"].map("{:,.2f}".format)
    # dfExport_pivot["Thành tiền (VND)"] = dfExport_pivot["Thành tiền (VND)"].map("{:,.0f}".format)

    dataframe_to_sheet(dfExport_pivot, Worksheet, row_start, col_start)

    # %% Edit table pivot
    #
    sheet_start = get_column_letter(col_start + 1) + str(row_start)
    sheet_end = get_column_letter(Worksheet.max_column) + str(row_start + 1 + len(dfExport_pivot))

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )
    set_border(Worksheet, f"{sheet_start}:{sheet_end}", thin_border)
    for row in range(row_start, len(dfExport_pivot) + row_start + 1):
        if str(Worksheet["E" + str(row)].value) == "nan":
            for i in range(1, Worksheet.max_column + 1):
                header = get_column_letter(i) + str(row)
                Worksheet[header].font = Font(bold=True)

    for i in range(1, Worksheet.max_column + 1):
        header = get_column_letter(i) + str(row_start)
        Worksheet[header].font = Font(bold=True)
        Worksheet.column_dimensions[get_column_letter(i)].bestFit = True

    if str.find(file_name, "NVTT") == -1:
        SumSL = f"{get_column_letter(Worksheet.max_column - 1)}{row_start+len(dfExport_pivot)+1}"
        Worksheet[SumSL] = TotalSL
        Worksheet[SumSL].font = Font(bold=True)
        SumTT = f"{get_column_letter(Worksheet.max_column)}{row_start+len(dfExport_pivot)+1}"
        Worksheet[SumTT] = TotalTT
        Worksheet[SumTT].font = Font(bold=True)
        Worksheet[f"B{row_start+len(dfExport_pivot)+1}"] = "Tổng cộng"

    for i in range(Worksheet.max_column - 1, Worksheet.max_column + 1):
        for j in range(1, len(dfExport_pivot) + row_start + 2):
            cell = get_column_letter(i) + str(j)
            Worksheet[cell].alignment = Alignment(horizontal="right")
            Worksheet[cell].number_format = "#,##0"

    Workbook.save(file_name)
    # .alignment = Alignment(horizontal='left')
    print(f"Save BCdonHang {file_name} => {sheet_name}")


# %%
# GetData from reportFolder
for f in os.listdir(reportFolder):
    if f.find("Bao_cao_chi_tiet_VTKHC2") > -1:
        dfVTKHC2 = pd.read_excel(os.path.join(reportFolder, f), engine="openpyxl")
        dfVTKHC2 = dfVTKHC2[3:].reset_index().drop(columns={"index"})
        dfVTKHC2.columns = dfVTKHC2.loc[0]
        dfVTKHC2 = dfVTKHC2.loc[1:]
        dfVTKHC2 = dfVTKHC2.drop(columns="STT")
        # dfVTKHC2.to_csv("dfVTKHC2.csv", index=False)

    if f.find("Bao_cao_du_lieu_don_hang") > -1:
        dfDuLieuDonHang = pd.read_excel(os.path.join(reportFolder, f), engine="openpyxl")
        dfDuLieuDonHang = dfDuLieuDonHang[3:].reset_index().drop(columns={"index"})
        dfDuLieuDonHang.columns = dfDuLieuDonHang.loc[0]
        dfDuLieuDonHang = dfDuLieuDonHang.loc[1:]
        dfDuLieuDonHang = dfDuLieuDonHang.drop(columns="STT")
        # dfDuLieuDonHang.to_csv("dfDuLieuDonHang.csv", index=False)


# %%
# Báo cáo Ghé thăm khách hàng C2 của NVTT
dfVTKHC2.columns = dfVTKHC2.columns.str.strip()
# Filter vùng đào tạo
filter = dfVTKHC2.loc[:, "Miền"].str.contains("đào tạo", case=False)
dfVTKHC2 = dfVTKHC2[~filter]
# Danh sách NVTT ghé thăm C2 trong tuần
dsMienGheTham = list(dfVTKHC2.loc[:, ["Miền"]].drop_duplicates().reset_index().drop(columns="index")["Miền"])

for Mien in dsMienGheTham:
    filter = dfVTKHC2["Miền"] == Mien
    folderGSBH = f"{reportFolder}/{Mien}/GSBH"

    try:
        os.mkdir(f"{reportFolder}/{Mien}")
    except Exception as e:
        print(e)
        pass

    try:
        os.mkdir(f"{reportFolder}/{Mien}/GSBH")
    except Exception as e:
        print(e)
        pass

    file_name = os.path.abspath(os.path.join(folderGSBH, "GSBH_" + Mien + ".xlsx"))
    # Xuất báo cáo chi tiết
    sheet_name = "BaocaoGheThamC2"
    dfExport = dfVTKHC2[filter]
    print(f"file_name {file_name}\n")

    BCGheThamC2(dfExport, file_name, sheet_name, reportFolder)


DSNVTT = list(dfVTKHC2.loc[:, ["MÃ NVTT"]].drop_duplicates().reset_index().drop(columns="index")["MÃ NVTT"])

for NVTT in DSNVTT:
    filter = dfVTKHC2["MÃ NVTT"] == NVTT
    dfExport = dfVTKHC2[filter]
    Mien = dfExport["Miền"].loc[dfExport.index[0]]
    folderNVTT = f"{reportFolder}/{Mien}/NVTT"
    try:
        os.mkdir(f"{reportFolder}/{Mien}")
    except Exception as e:
        pass

    try:
        os.mkdir(f"{reportFolder}/{Mien}/NVTT")
    except Exception as e:
        pass
    file_name = os.path.abspath(os.path.join(folderNVTT, "NVTT_" + NVTT + ".xlsx"))
    # Xuất báo cáo chi tiết
    sheet_name = "BaocaoGheThamC2"

    BCGheThamC2(dfExport, file_name, sheet_name, reportFolder)

# print(f"{file_name} - {sheet_name}")


# %%
# Báo cáo đơn hàng C2
# Remove leading/trailing whitespaces from column names
dfDuLieuDonHang.columns = dfDuLieuDonHang.columns.str.strip()
# Filter rows where the 'Khu vực' column contains 'đào tạo'
filter = dfDuLieuDonHang.loc[:, "Khu vực"].str.contains("đào tạo", case=False)
dfDuLieuDonHang = dfDuLieuDonHang[~filter]
# Create a list of unique 'Miền' values
dsMienDonHang = list(dfDuLieuDonHang.loc[:, ["Miền"]].drop_duplicates().reset_index().drop(columns="index")["Miền"])
folderGSBH = os.path.join(reportFolder, "GSBH")
for Mien in dsMienDonHang:
    # Xuất báo cáo chi tiết
    filter = dfDuLieuDonHang["Miền"] == Mien
    dfExport = dfDuLieuDonHang[filter]
    try:
        os.mkdir(f"{reportFolder}/{Mien}")
    except Exception as e:
        pass

    try:
        os.mkdir(f"{reportFolder}/{Mien}/GSBH")
    except Exception as e:
        pass
    folderGSBH = f"{reportFolder}/{Mien}/GSBH"

    file_name = os.path.abspath(os.path.join(folderGSBH, "GSBH_" + Mien + ".xlsx"))
    sheet_name = "BaocaoDonHangC2"
    BCDonHangC2(dfExport, file_name, sheet_name, reportFolder)


dsNVTT = list(dfDuLieuDonHang.loc[:, ["Mã NVTT"]].drop_duplicates().reset_index().drop(columns="index")["Mã NVTT"])
folderNVTT = os.path.join(reportFolder, "NVTT")
for NVTT in dsNVTT:
    filter = dfDuLieuDonHang["Mã NVTT"] == NVTT
    dfExport = dfDuLieuDonHang[filter]
    Mien = dfExport["Miền"].loc[dfExport.index[0]]
    try:
        os.mkdir(f"{reportFolder}/{Mien}")
    except Exception as e:
        pass

    try:
        os.mkdir(f"{reportFolder}/{Mien}/NVTT")
    except Exception as e:
        pass
    folderNVTT = f"{reportFolder}/{Mien}/NVTT"

    file_name = os.path.abspath(os.path.join(folderNVTT, "NVTT_" + NVTT + ".xlsx"))
    sheet_name = "BaocaoDonHangC2"

    # if len(dfExport) >0:
    BCDonHangC2(dfExport, file_name, sheet_name, reportFolder)
